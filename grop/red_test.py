from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

wb = Workbook()
ws = wb.active

ws['D2'] = 2250 # 2000以上

cell = ws['D2']
cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
cell.font = Font(color="FFFFFF")

wb.save("赤テスト2.xlsx")
print("赤テスト2.xlsx ができた!D2が赤くなってる確認して")