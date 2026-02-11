# Day 9: フォルダを指定して複数ファイルをまとめる

print("=== Day 9: フォルダを指定して集計します! ===")

import pandas as pd
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# 1. フォルダパスをユーザーに聞く
folder_path = input("まとめるファイルがあるフォルダのパスを入力してください").strip()

# 入力が空なら現在のフォルダを使う
if not folder_path:
    folder_path = os.getcwd()
    print("フォルダパスが空だったので、現在のフォルダを使います")

# フォルダが存在するかチェック
if not os.path.isdir(folder_path):
    print(f"フォルダが見つかりません: {folder_path}")
    exit()

print(f"指定フォルダ: {folder_path}")

# 2. 指定フォルダ内のxlsxファイルを全部探す
excel_files = glob.glob(os.path.join(folder_path, "*.xlsx"))

print(f"見つかったファイル数: {len(excel_files)}")
if len(excel_files) == 0:
    print("xlsxファイルが見つかりませんでした")
    exit()

for f in excel_files:
    print(" - " + os.path.basename(f))

# 3. 全部読み込んでまとめる
all_dfs = []
for file in excel_files:
    try:
        df = pd.read_excel(file)
        all_dfs.append(df)
    except Exception as e:
        print(f"エラー(スキップ): {os.path.basename(file)} → {e}")

if all_dfs:
    combined_df = pd.concat(all_dfs, ignore_index=True)
    combined_df["合計金額"] = combined_df["単価"] * combined_df["数量"]

    # 合計行(単価「-」)
    total_row = pd.DataFrame({
        "商品": ["合計"],
        "単価": [" - "],
        "数量": [combined_df["数量"].sum()],
        "合計金額": [combined_df["合計金額"].sum()]
    })
    combined_df = pd.concat([combined_df, total_row], ignore_index=True)

    print("\nまとめた表(最初の5行):")
    print(combined_df.head())

    # 4.出力保存 + 書式付け(Day 8の修正版を使用)
    output_file = os.path.join(folder_path, "集計結果_フォルダ指定.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "集計結果"

    # ヘッダー
    headers = list(combined_df.columns)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignrment(horizontal="center")

    #データ
    for row_idx, row in combined_df.iterrows():
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx+2, column=col, value=value)

            if col == 4: #合計金額列
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            if row_idx + 2 == len(combined_df) + 1: #合計行
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.font = Font(bold=True)
                cell.border = Border(bottom=Side(style="medium"))

    # 列幅調整
    column_widths = [15, 10, 10, 15]
    for col_num, width in enumerate(column_widths, 1):
        column_letter = chr(64 + col_num)
        ws.column_dimensions[column_letter].width = width

    wb.save(output_file)

    print(f"\n保存完了!場所: {output_file}")
    print("フォルダ内で開いて確認してね!")
else:
    print("読み込めるファイルがありませんでした")