# Day10: 条件付き色付け(合計金額2000以上の行を赤く)

print("=== Day10: 条件付き色付けを追加します! ===")

import pandas as pd
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# 条件: 合計金額がこの値以上の行を赤くする
RED_THRESHOLD = 2000 #←ここを変えれば調整可能

# 1. フォルダ指定(Day 9と同じ)
folder_path = input("まとめるファイルがあるフォルダのパスを入力してください").strip()
# C:\Users\PC\PythonProjects\MyProject\grop\売上データ      
if not folder_path:
    folder_path = os.getcwd()

folder_path = folder_path.strip('"') #引用符除去

if not os.path.isdir(folder_path):
    print(f"フォルダが見つかりません: {folder_path}")
    exit()

excel_files = glob.glob(os.path.join(folder_path, "*.xlsx"))

print(f"見つかったファイル数: {len(excel_files)}")

all_dfs = []
for file in excel_files:
    try:
        df = pd.read_excel(file)
        # 数値列を強制変換(これで文字列問題解決)
        df["単価"] = pd.to_numeric(df["単価"], errors='coerce')
        df["数量"] = pd.to_numeric(df["数量"], errors='coerce')
        df["合計金額"] = df["単価"] * df["数量"]
        all_dfs.append(df)
    except:
        pass

if all_dfs:
    combined_df = pd.concat(all_dfs, ignore_index=True)

    total_row = pd.DataFrame({
        "商品": ["合計"],
        "単価": ["-"],
        "数量": [combined_df["数量"].sum()],
        "合計金額": [combined_df["合計金額"].sum()]
    })
    combined_df = pd.concat([combined_df, total_row], ignore_index=True)

    # 出力 + 書式付け
    output_file = os.path.join(folder_path, "集計結果_条件修正.xlsx")

    wb =Workbook()
    ws = wb.active
    ws.title = "集計結果"

    #ヘッダー
    headers = list(combined_df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # データ書き込み + 条件色付け
    last_row = len(combined_df) + 1 #合計行の行番号
    for r, row in enumerate(combined_df.itertuples(index=False), start=2):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=value)

            
        # 条件: 合計金額が闘値以上→赤背景(合計行は除く)
            if c == 4:
                try:
                    num_value = float(value)
                    if num_value >= RED_THRESHOLD:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True) # 白文字で読みやすく
                    # 合計金額列(D列)を黄色
                    else:
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                except:
                    pass

            # 合計行(最後)をグレー + 太字
            if r == last_row:
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.font = Font(bold=True)
                cell.border = Border(bottom=Side(style="medium"))
                continue
    # 列幅調整
    column_widths = [15, 10, 10, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width

    wb.save(output_file)

    print(f"\n保存完了!場所: {output_file}")
    print(f"合計金額{RED_THRESHOLD}円以上の行が赤くなってるはず!")
    print("開いて確認してね")
else:
    print("ファイルなし")