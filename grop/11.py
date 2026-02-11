# Excel自動集計ツール(完成版)
# 機能:
# 指定ファイル内のXlsxファイルを全部読み込み
# データを一つにまとめて合計金額計算
# 合計行追加
# 書式付け(ヘッダー文字、合計金額黄色、合計行グレー)
# 条件色付け(合計金額2000円以上を赤く)
# 出力ファイルに保存

print("=== Excel自動集計ツール(完成版) ===")

import pandas as pd
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# 設定(ここを変えれば調整可能)
RED_THRESHOLD = 2000 #赤くする合計金額の闘値
OUTPUT_FILENAME = "自動集計結果.xlsx" # 出力ファイル名

# 1. フォルダパス入力
folder_path = input("集計したいフォルダのパスを入力してください")
# C:\Users\PC\PythonProjects\MyProject\grop\売上データ 
if not folder_path:
    folder_path = os.getcwd()
    print("パスが空だったので現在のフォルダを使います")

# バッグスラッシュを正しく扱う
folder_path = folder_path.replace('\\', '\\\\').strip('"')

if not os.path.isdir(folder_path):
    print(f"フォルダが見つかりません: {folder_path}")
    print("パスを確認して再実行してください")
    exit()

print(f"対象フォルダ: {folder_path}")

# 2. xlsxファイルを全部探す
excel_files = glob.glob(os.path.join(folder_path, "*.xlsx"))

if len(excel_files) == 0:
    print("xlsxファイルが見つかりませんでした")
    exit()

print(f"見つかったファイル: {len(excel_files)}個")
for f in excel_files:
    print(f" - {os.path.basename(f)}")

# 3. 全部読み込んで結合
all_dfs = []
for file in excel_files:
    try:
        df = pd.read_excel(file)
        # 数値変換(エラー防止)
        df["単価"] = pd.to_numeric(df["単価"], errors='coerce').fillna(0)
        df["数量"] = pd.to_numeric(df["数量"], errors='coerce').fillna(0)
        df["合計金額"] = df["単価"] * df["数量"]
        all_dfs.append(df)
    except Exception as e:
        print(f"読み込みエラー(スキップ): {os.path.basename(file)} → {e}")

if not all_dfs:
    print("有効なデータがありませんでした")
    exit()

combined_df = pd.concat(all_dfs, ignore_index=True)

# 合計行追加
total_row = pd.DataFrame({
    "商品": ["合計"],
    "単価": ["-"],
    "数量": [combined_df["数量"].sum()],
    "合計金額": [combined_df["合計金額"].sum()]
})
combined_df = pd.concat([combined_df, total_row], ignore_index=True)

print("\n集計結果!(最初の5行)")
print(combined_df.head())

# 4. openpyxlで出力 + 書式付け
output_file = os.path.join(folder_path, OUTPUT_FILENAME)

wb = Workbook()
ws = wb.active
ws.title = "集計結果"

#ヘッダー
headers = list(combined_df.columns)
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    
# データ書き込み + 条件色付け
last_row = len(combined_df) + 1 #合計行の行番号
for r, row in enumerate(combined_df.itertuples(index=False), start=2):
    for c, value in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=value)

            
         # 条件: 合計金額が闘値以上→赤背景(合計行は除く)
        if c == 4:
            try:
                num_value = float(value)
                if num_value >= RED_THRESHOLD:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True) # 白文字で読みやすく
                # 合計金額列(D列)を黄色
                else:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            except:
                pass

        # 合計行(最後)をグレー + 太字
        if r == last_row:
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style="medium"))
            continue

    # 列幅調整
    column_widths = [15, 10, 10, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width

    wb.save(output_file)

    print(f"\n保存完了!場所: {output_file}")
    print(f"合計金額{RED_THRESHOLD}円以上の行が赤くなってるはず!")
    print("開いて確認してね")
else:
    print("ファイルなし")
