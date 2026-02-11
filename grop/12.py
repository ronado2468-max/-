# Excel 自動集計ツール(スケジュール対応版)

import pandas as pd
import glob
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# 設定
RED_THRESHOLD = 2000
OUTPUT_FILENAME = "自動集計結果.xlsx"

def main(folder_path=None):
    if folder_path is None:
        if len(sys.argv) > 1:
            folder_path = sys.argv[1] # コマンドライン引数からパスを取る
        else:
            folder_path = input("集計フォルダのパスを入力").strip()
            if not folder_path:
                folder_path = os.getcwd()

    folder_path = folder_path.strip('"').replace('\\', '\\\\')

    if not os.path.isdir(folder_path):
        print(f"フォルダが見つかりません: {folder_path}")
        return

    excel_files = glob.glob(os.path.join(folder_path, "*.xlsx"))
    if not excel_files:
        print("xlsxファイルなし")
        return

    all_dfs = []
    for file in excel_files:
        try:
            df = pd.read_excel(file)
            df["単価"] = pd.to_numeric(df["単価"], errors='coerce').fillna(0)
            df["数量"] = pd.to_numeric(df["数量"], errors='coerce').fillna(0)
            df["合計金額"] = df["単価"] * df["数量"]
            all_dfs.append(df)
        except:
            pass

    if not all_dfs:
        return

    combined_df = pd.concat(all_dfs, ignore_index=True)

    total_row = pd.DataFrame({
        "商品": ["合計"],
        "単価": ["-"],
        "数量": [combined_df["数量"].sum()],
        "合計金額": [combined_df["合計金額"].sum()]
    })
    combined_df = pd.concat([combined_df, total_row], ignore_index=True)

    output_file = os.path.join(folder_path, OUTPUT_FILENAME)

    wb = Workbook()
    ws = wb.active

    headers = list(combined_df.columns)
    for col_num, header in enumerate(headers, 1):
       cell = ws.cell(row=1, column=col_num, value=header)
       cell.font = Font(bold=True)
       cell.alignment = Alignment(horizontal="center")

    last_row = len(combined_df) + 1
    for r, row in enumerate(combined_df.itertuples(index=False), start=2):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=value)

            if c == 4:
                try:
                    num_value = float(value)
                    if num_value >= RED_THRESHOLD:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    else:
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                except:
                    pass

            if r == last_row:
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.font = Font(bold=True)
                cell.border = Border(bottom=Side(style="medium"))
                continue

    column_widths = [15, 10, 10, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width

    wb.save(output_file)
    print(f"自動集計完了!出力: {output_file}")

if __name__ == "__main__":
    main()

    #C:\Users\PC\AppData\Local\Programs\Python\Python314\python.exe
    #f