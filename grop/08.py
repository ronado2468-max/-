# Day 8: 複数Excelファイルをまとめて + 出力ファイルに書式付け

print("=== Day 8: 出力ファイルをキレイにします! ===")

import pandas as pd
import glob # ファイル一覧を取るためのライブラリ
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# 1. 対象のExcelファイルを全部探す(同じフォルダ内の .xlsx 全部)
# excel_files = glob.glob(r"C:\Users\PC\PythonProjects\MyProject\grop\*.xlsx") # *.xlsxで全部のExcelを探す

# もし特定のファイルだけならこう
excel_files = glob.glob(r"C:\Users\PC\PythonProjects\MyProject\grop\集計結果.xlsx") # salesで始まるファイルだけ

all_dfs = []
for file in excel_files:
    try:
        df = pd.read_excel(file)
        all_dfs.append(df)
    except:
        pass
    
if all_dfs:
    combined_df = pd.concat(all_dfs, ignore_index=True)

    #合計金額追加
    combined_df["合計金額"] = combined_df["単価"] * combined_df["数量"]

    #全体合計行を作成
    total_row = pd.DataFrame({
        "商品": ["合計"],
        "単価": ["-"],
        "数量": [combined_df["数量"].sum()],
        "合計金額": [combined_df["合計金額"].sum()]
    })

    #合計行を追加
    combined_df = pd.concat([combined_df, total_row], ignore_index=True)

    print("\nまとめた表(書式前):")
    print(combined_df)

    
# 2. openpyxlで新しいエクセルを作って書式付け
    output_file = r"C:\Users\PC\PythonProjects\MyProject\grop\集計結果.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "集計結果"

    #ヘッダー読み込み(太字に)
    headers = list(combined_df.columns)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    #データ書き込み
    for row_idx, row in combined_df.iterrows():
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx+2, column=col, value=value)

            #合計金額(D列)を黄色背景に
            if col == 4: #D列 = 合計金額
                cell.fill = PatternFill(
                    start_color="FFFF00",
                    end_color="FFFF00", 
                    fill_type="solid"
                )
    
    #合計行(最後)を太字 + 下線
    last_row = ws.max_row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=last_row,column=col)
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(style="thin"))
    
    #列幅自動調整
    #for col in ws.columns:
        #max_length = 0
        #column = col[0].column_letter
        #for cell in col:
            #if cell.value:
                #max_length = max(max_length, len(str(cell.value)))
        #ws.column_dimensions[column].width = max_length + 2

    #列幅を強制調整(日本語対応強化)
    column_widths = [15, 10, 10, 15] # 商品, 単価, 数量, 合計金額の幅
    for col_num, width in enumerate(column_widths, 1):
        column_letter = chr(64 + col_num) # A=1. B=2
        ws.column_dimensions[column_letter].width = width

    #保存
    wb.save(output_file)

    print(f"\n保存完了!ファイル: {output_file}")
    print("聞いてみて!合計行が太字で、合計金額列が黄色になってるはず")
else:
    print("ファイルが見つかりません")


